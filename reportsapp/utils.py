"""Shared helpers for generating PDF and Excel reports across dashboards."""
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from functools import lru_cache
from django.conf import settings
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib.utils import ImageReader
from fees.utils import format_ugx

SCHOOL_BADGE_PATH = settings.BASE_DIR / "static" / "img" / "school_badge.jpeg"


@lru_cache(maxsize=1)
def _watermark_image_reader():
    if not SCHOOL_BADGE_PATH.exists():
        return None
    badge = PILImage.open(SCHOOL_BADGE_PATH).convert("RGBA")
    faded = PILImage.new("RGBA", badge.size, (255, 255, 255, 0))
    faded = PILImage.blend(faded, badge, alpha=0.08)
    buffer = io.BytesIO()
    faded.save(buffer, format="PNG")
    buffer.seek(0)
    return ImageReader(buffer)


def _draw_watermark(canvas_obj, doc):
    watermark = _watermark_image_reader()
    if watermark is None:
        return
    page_width, page_height = doc.pagesize
    size = min(page_width, page_height) * 0.6
    canvas_obj.saveState()
    canvas_obj.drawImage(
        watermark,
        (page_width - size) / 2,
        (page_height - size) / 2,
        width=size,
        height=size,
        mask="auto",
        preserveAspectRatio=True,
    )
    canvas_obj.restoreState()


@lru_cache(maxsize=1)
def _excel_logo_path():
    if not SCHOOL_BADGE_PATH.exists():
        return None
    resized_path = SCHOOL_BADGE_PATH.with_name("school_badge_logo_small.png")
    if not resized_path.exists():
        badge = PILImage.open(SCHOOL_BADGE_PATH).convert("RGBA")
        badge.thumbnail((90, 135))
        badge.save(resized_path, format="PNG")
    return str(resized_path)


def export_excel(filename, title, headers, rows):
    """
    headers: list[str]
    rows: list[list] (values only, in same order as headers)
    Returns an HttpResponse with an .xlsx attachment.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    header_row_idx = 3
    logo_path = _excel_logo_path()
    if logo_path:
        img = XLImage(logo_path)
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 80
        title_col_start = 3
    else:
        title_col_start = 1

    ws.merge_cells(
        start_row=1, start_column=title_col_start, end_row=1, end_column=max(len(headers), title_col_start)
    )
    title_cell = ws.cell(row=1, column=title_col_start, value=title)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_offset, row in enumerate(rows, start=header_row_idx + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_offset, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows] or [10])
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_pdf(filename, title, headers, rows, subtitle=None, landscape_mode=False):
    """
    headers: list[str]
    rows: list[list] (values only, in same order as headers)
    Returns an HttpResponse with a .pdf attachment.
    """
    buffer = io.BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer, pagesize=pagesize,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    data = [headers] + [[str(v) for v in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response

def export_report_card_pdf(filename, student, term_label, year, subject_rows, total, average, position=None, class_size=None, fee_status=None):
    """
    Build a single-pupil report card PDF (as opposed to export_pdf, which is a
    generic multi-row table export). subject_rows: list of (subject_name, score, grade).
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lyantonde Model Primary School", styles["Title"]),
        Paragraph(f"Pupil Report Card &mdash; {term_label} {year}", styles["Heading2"]),
        Spacer(1, 12),
    ]

    info_data = [
        ["Name:", student.full_name, "Admission No.:", student.admission_number],
        ["Class:", str(student.school_class) if student.school_class else "-", "Gender:", student.get_gender_display()],
        ["Boarding Status:", student.get_boarding_status_display(), "", ""],
    ]
    info_table = Table(info_data, colWidths=[3 * cm, 6 * cm, 3.5 * cm, 4 * cm])
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(info_table)
    elements.append(Spacer(1, 16))

    subject_data = [["Subject", "Score", "Grade"]] + [
        [name, f"{score:.1f}", grade] for name, score, grade in subject_rows
    ]
    subject_table = Table(subject_data, colWidths=[8 * cm, 4 * cm, 4 * cm], repeatRows=1)
    subject_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(subject_table)
    elements.append(Spacer(1, 16))

    summary_lines = [f"<b>Total:</b> {total:.1f}", f"<b>Average:</b> {average:.1f}" if average is not None else "<b>Average:</b> -"]
    if position and class_size:
        summary_lines.append(f"<b>Position:</b> {position} out of {class_size}")
    elements.append(Paragraph(" &nbsp;&nbsp;|&nbsp;&nbsp; ".join(summary_lines), styles["Normal"]))

    if fee_status is not None:
        elements.append(Spacer(1, 10))
        fee_line = (
            f"<b>Fees ({term_label} {year}):</b> Expected {format_ugx(fee_status['expected'])}, "
            f"Paid {format_ugx(fee_status['paid'])}, Balance {format_ugx(fee_status['balance'])} "
            f"({'DEFAULTER' if fee_status['is_defaulter'] else 'Cleared'})"
        )
        elements.append(Paragraph(fee_line, styles["Normal"]))

    elements.append(Spacer(1, 40))
    signature_data = [["Class Teacher's Signature: ______________________", "Headteacher's Signature: ______________________"]]
    signature_table = Table(signature_data, colWidths=[9 * cm, 9 * cm])
    signature_table.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 10)]))
    elements.append(signature_table)

    doc.build(elements, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export_nursery_report_card_pdf(
    filename, student, term_label, year, nursery_rows, position=None, class_size=None, fee_status=None
):
    """Build the compact A4 report used by Nursery, Middle, and Top classes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    centered = styles["Normal"].clone("NurseryCentered")
    centered.alignment = TA_CENTER
    centered.fontSize = 9
    centered.leading = 12

    elements = []
    logo = str(SCHOOL_BADGE_PATH) if SCHOOL_BADGE_PATH.exists() else None
    heading = [
        Paragraph("<b>LYANTONDE MODEL PRIMARY SCHOOL</b>", styles["Title"]),
        Paragraph("P.O. BOX 93, LYANTONDE", centered),
        Paragraph("0756001495 / 0752834565 / 0789228711", centered),
        Paragraph('<i>Motto: "We strive for quality education"</i>', centered),
        Paragraph("<b>END OF TERM REPORT</b>", centered),
    ]
    if logo:
        from reportlab.platypus import Image

        header = Table([[Image(logo, width=2 * cm, height=2 * cm), heading]], colWidths=[2.5 * cm, 14.5 * cm])
        header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elements.append(header)
    else:
        elements.extend(heading)
    elements.append(Spacer(1, 8))

    pupil_data = [
        ["PUPIL'S NAME", student.full_name, "SEX", student.get_gender_display()],
        ["CLASS", str(student.school_class), "TERM", term_label],
        ["POSITION", f"{position or '-'} out of {class_size or '-'}", "YEAR", str(year)],
    ]
    pupil_table = Table(pupil_data, colWidths=[3 * cm, 7 * cm, 2.5 * cm, 4.5 * cm])
    pupil_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (1, 0), (1, -1), 0.35, colors.grey),
        ("LINEBELOW", (3, 0), (3, -1), 0.35, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.extend([pupil_table, Spacer(1, 10), Paragraph("<b>END OF TERM RESULTS</b>", centered)])

    result_data = [["SUBJECT", "MARKS (%)", "REMARKS", "INITIALS"]]
    result_data.extend([
        [
            row["subject"],
            f'{row["score"]:.0f}' if row["score"] is not None else "-",
            row["remarks"] or "-",
            row["initials"] or "-",
        ]
        for row in nursery_rows
    ])
    if not nursery_rows:
        result_data.append(["No results recorded", "-", "-", "-"])
    total_score = sum(row["score"] for row in nursery_rows if row["score"] is not None)
    result_data.append(["TOTAL", f"{total_score:.0f}", "", ""])
    results_table = Table(result_data, colWidths=[6 * cm, 3 * cm, 5.5 * cm, 2.5 * cm], repeatRows=1)
    results_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7ECE8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.extend([results_table, Spacer(1, 14)])

    elements.append(Paragraph("Class Teacher's comment: _________________________________________________", styles["Normal"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Headteacher's report: ____________________________________________________", styles["Normal"]))
    elements.append(Spacer(1, 18))
    if fee_status is not None:
        elements.append(Paragraph(
            f"Next term begins on: ____________________ &nbsp;&nbsp;&nbsp; "
            f"School fees balance: <b>{format_ugx(fee_status['balance'])}</b>",
            styles["Normal"],
        ))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph("I have seen and read the report. Parent/Guardian: __________________________", styles["Normal"]))

    doc.build(elements, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export_progressive_report_card_pdf(
    filename, student, term_label, year, rows, division, fee_status=None
):
    """Build the P4-P7 progressive report using the school's paper layout."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, leftMargin=1.1 * cm, rightMargin=1.1 * cm,
        topMargin=1.1 * cm, bottomMargin=1.1 * cm,
    )
    styles = getSampleStyleSheet()
    centered = styles["Normal"].clone("ProgressiveCentered")
    centered.alignment = TA_CENTER
    centered.fontSize = 9
    centered.leading = 12

    heading = [
        Paragraph("<b>LYANTONDE MODEL PRIMARY SCHOOL</b>", styles["Title"]),
        Paragraph("P.O. BOX 93, LYANTONDE", centered),
        Paragraph("0756001495 / 0765782480 / 0789228711", centered),
        Paragraph('<i>Motto: "We strive for quality education"</i>', centered),
        Paragraph("<b><u>PROGRESSIVE REPORT</u></b>", centered),
    ]
    elements = []
    if SCHOOL_BADGE_PATH.exists():
        from reportlab.platypus import Image

        header = Table(
            [[Image(str(SCHOOL_BADGE_PATH), width=2.2 * cm, height=2.2 * cm), heading]],
            colWidths=[2.7 * cm, 14.2 * cm],
        )
        header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elements.append(header)
    else:
        elements.extend(heading)
    elements.append(Spacer(1, 8))

    details = Table([
        ["PUPIL'S NAME", student.full_name, "SEX", student.get_gender_display()],
        ["CLASS", str(student.school_class), "TERM", term_label, "YEAR", str(year)],
    ], colWidths=[2.6 * cm, 7.2 * cm, 1.5 * cm, 2.2 * cm, 1.4 * cm, 2 * cm])
    details.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTNAME", (4, 1), (4, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (1, 0), (1, -1), 0.35, colors.grey),
        ("LINEBELOW", (3, 0), (3, -1), 0.35, colors.grey),
        ("LINEBELOW", (5, 1), (5, 1), 0.35, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.extend([details, Spacer(1, 7)])

    headers = ["SUBJECT", "SET\n1", "SET\n2", "SET\n3", "SET\n4", "MID", "SET\n6", "SET\n7", "SET\n8", "SET\n9", "END", "TOTAL", "AVE", "AGG", "RMKS", "INITI"]
    table_data = [headers]
    for row in rows:
        score = f'{row["score"]:.0f}' if row["score"] is not None else ""
        table_data.append([
            row["subject"], "", "", "", "", "", "", "", "", "", score,
            score, score, row["aggregate"] or "", row["remarks"], row["initials"],
        ])
    entered_scores = [row["score"] for row in rows if row["score"] is not None]
    total_score = sum(entered_scores)
    average = total_score / len(entered_scores) if entered_scores else None
    aggregate = sum(row["aggregate"] for row in rows if row["aggregate"] is not None)
    table_data.append([
        "TOTAL", "", "", "", "", "", "", "", "", "", "",
        f"{total_score:.0f}" if entered_scores else "",
        f"{average:.1f}" if average is not None else "",
        aggregate or "", "", "",
    ])
    widths = [1.55 * cm] + [0.72 * cm] * 10 + [0.85 * cm, 0.8 * cm, 0.8 * cm, 1.35 * cm, 0.8 * cm]
    marks_table = Table(table_data, colWidths=widths, repeatRows=1)
    marks_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7ECE8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.extend([marks_table, Spacer(1, 16)])
    elements.append(Paragraph(f"<b><u>DIVISION</u>:</b> {division}", styles["Normal"]))
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("Class teacher's comment: _________________________________________________", styles["Normal"]))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Head teacher's report: ___________________________________________________", styles["Normal"]))
    elements.append(Spacer(1, 22))
    elements.append(Paragraph("<b><u>REQUIREMENTS</u></b>", styles["Normal"]))
    elements.append(Spacer(1, 7))
    elements.append(Paragraph(
        "A ream of paper, 4 toilet papers, 2 brooms, a bucket of detergent, enough books, "
        "enough pens, mathematical set and atlas.", styles["Normal"]
    ))
    elements.append(Spacer(1, 12))
    if fee_status is not None:
        elements.append(Paragraph(
            f"Next term begins on ____________________ &nbsp;&nbsp; School fees balance: "
            f"<b>{format_ugx(fee_status['balance'])}</b>", styles["Normal"]
        ))
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("I have seen and read the report, sign __________________ Parent/guardian", styles["Normal"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Date __________________", styles["Normal"]))

    doc.build(elements, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response
