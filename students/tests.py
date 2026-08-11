from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook
import io

from accounts.models import User

from .models import Requirement, SchoolClass, StudentRequirement
from .forms import StudentForm


class ClassManagementTests(TestCase):
    def setUp(self):
        self.headteacher = User.objects.create_user(
            username="headteacher", password="test-password", role=User.Role.HEADTEACHER
        )
        self.client.force_login(self.headteacher)

    def test_add_class(self):
        response = self.client.post(
            reverse("students:manage_classes"), {"name": "Middle Class", "class_teacher": ""}
        )

        self.assertRedirects(response, reverse("students:manage_classes"))
        self.assertTrue(SchoolClass.objects.filter(name="Middle Class").exists())

    def test_duplicate_class_name_displays_validation_error(self):
        SchoolClass.objects.create(name="Top Class")

        response = self.client.post(
            reverse("students:manage_classes"), {"name": "Top Class", "class_teacher": ""}
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "School class with this Name already exists")

    def test_edit_class_page_loads_and_saves(self):
        school_class = SchoolClass.objects.create(name="Nursery")
        url = reverse("students:edit_class", args=[school_class.pk])

        self.assertEqual(self.client.get(url).status_code, 200)
        response = self.client.post(url, {"name": "Baby Class", "class_teacher": ""})

        self.assertRedirects(response, reverse("students:manage_classes"))
        school_class.refresh_from_db()
        self.assertEqual(school_class.name, "Baby Class")

    def test_headteacher_creates_staff_member_who_appears_in_class_form(self):
        response = self.client.post(reverse("students:manage_staff"), {
            "first_name": "Grace",
            "last_name": "Teacher",
            "username": "grace.teacher",
            "email": "grace@example.com",
            "phone_number": "0700000000",
            "password1": "StrongTestPassword123!",
            "password2": "StrongTestPassword123!",
        })

        self.assertRedirects(response, reverse("students:manage_staff"))
        teacher = User.objects.get(username="grace.teacher")
        self.assertEqual(teacher.role, User.Role.TEACHER)
        class_page = self.client.get(reverse("students:manage_classes"))
        self.assertContains(class_page, 'value="{}"'.format(teacher.pk))

    def test_inactive_teacher_is_not_available_for_class_assignment(self):
        teacher = User.objects.create_user(
            username="inactive-teacher", password="test-password", role=User.Role.TEACHER, is_active=False
        )

        response = self.client.get(reverse("students:manage_classes"))

        self.assertNotContains(response, 'value="{}"'.format(teacher.pk))


class UgandanNinValidationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school_class = SchoolClass.objects.create(name="NIN Test Class")

    def student_data(self, nin):
        return {
            "first_name": "Test",
            "last_name": "Pupil",
            "gender": "F",
            "date_of_birth": "2015-01-01",
            "school_class": self.school_class.pk,
            "boarding_status": "DAY",
            "former_school": "",
            "religion": "",
            "nin": nin,
            "guardian_name": "Test Guardian",
            "guardian_phone": "0700000000",
            "address": "Lyantonde",
            "date_admitted": "2026-01-01",
        }

    def test_accepts_and_normalizes_valid_ugandan_nin(self):
        form = StudentForm(data=self.student_data("cm4900906p76ze"))

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["nin"], "CM4900906P76ZE")

    def test_rejects_wrong_prefix_or_length(self):
        for nin in ("UG4900906P76ZE", "CM123", "CM4900906P76ZE9", "CF4900906P76-_"):
            with self.subTest(nin=nin):
                form = StudentForm(data=self.student_data(nin))
                self.assertFalse(form.is_valid())
                self.assertIn("nin", form.errors)

    def test_nin_remains_optional(self):
        form = StudentForm(data=self.student_data(""))

        self.assertTrue(form.is_valid(), form.errors)


class RequirementsRegisterTests(TestCase):
    def setUp(self):
        self.headteacher = User.objects.create_user(
            username="requirements-head", password="test-password", role=User.Role.HEADTEACHER
        )
        self.school_class = SchoolClass.objects.create(name="P6")
        self.student = self.school_class.students.create(
            first_name="Requirements",
            last_name="Pupil",
            gender="M",
            date_of_birth="2015-01-01",
            guardian_name="Parent",
            guardian_phone="0700000000",
        )
        self.boarding_student = self.school_class.students.create(
            first_name="Boarding",
            last_name="Pupil",
            gender="F",
            date_of_birth="2015-02-01",
            boarding_status="BOARDING",
            guardian_name="Parent",
            guardian_phone="0700000001",
        )
        self.client.force_login(self.headteacher)

    def test_register_loads_requirements_for_selected_class_group(self):
        response = self.client.get(reverse("students:requirements_register"), {
            "class_id": self.school_class.pk, "term": "TERM1", "year": 2026,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "1 mathematics set")
        self.assertContains(response, "Atlas")
        self.assertNotContains(response, "1 packet of crayons")

    def test_saves_checked_and_unchecked_requirements(self):
        requirements = list(Requirement.objects.filter(
            class_group=Requirement.ClassGroup.P4_P7, scholar_type="DAY"
        ))
        checked = requirements[:2]
        response = self.client.post(reverse("students:requirements_register"), {
            "class_id": self.school_class.pk,
            "term": "TERM2",
            "year": 2026,
            "brought": [f"{self.student.pk}:{requirement.pk}" for requirement in checked],
        })

        self.assertEqual(response.status_code, 302)
        records = StudentRequirement.objects.filter(student=self.student, term="TERM2", year=2026)
        self.assertEqual(records.count(), len(requirements))
        self.assertEqual(records.filter(brought=True).count(), 2)
        self.assertTrue(all(record.brought_on is not None for record in records.filter(brought=True)))
        self.assertTrue(all(record.recorded_by == self.headteacher for record in records))

    def test_downloads_saved_requirements_list(self):
        requirement = Requirement.objects.filter(
            class_group=Requirement.ClassGroup.P4_P7, scholar_type="DAY"
        ).first()
        StudentRequirement.objects.create(
            student=self.student,
            requirement=requirement,
            term="TERM1",
            year=2026,
            brought=True,
            recorded_by=self.headteacher,
        )

        response = self.client.get(reverse("students:export_requirements_register"), {
            "class_id": self.school_class.pk,
            "term": "TERM1",
            "year": 2026,
            "scholar_type": "DAY",
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        values = list(workbook.active.iter_rows(values_only=True))
        flattened = {value for row in values for value in row if value is not None}
        self.assertIn(self.student.admission_number, flattened)
        self.assertIn("Brought", flattened)
        self.assertIn("Missing", flattened)

    def test_boarding_register_is_separate_from_day_register(self):
        boarding_response = self.client.get(reverse("students:requirements_register"), {
            "class_id": self.school_class.pk,
            "term": "TERM1",
            "year": 2026,
            "scholar_type": "BOARDING",
        })
        day_response = self.client.get(reverse("students:requirements_register"), {
            "class_id": self.school_class.pk,
            "term": "TERM1",
            "year": 2026,
            "scholar_type": "DAY",
        })

        self.assertContains(boarding_response, self.boarding_student.full_name)
        self.assertContains(boarding_response, 'name="brought"')
        self.assertContains(boarding_response, "requirement-select-all")
        self.assertNotContains(boarding_response, self.student.full_name)
        self.assertContains(boarding_response, "Mattress, blanket")
        self.assertNotContains(boarding_response, "4 toilet rolls")
        self.assertContains(day_response, self.student.full_name)
        self.assertNotContains(day_response, self.boarding_student.full_name)
        self.assertContains(day_response, "4 toilet rolls")
        self.assertNotContains(day_response, "Mattress, blanket")

    def test_teacher_can_record_requirements_for_assigned_class(self):
        teacher = User.objects.create_user(
            username="class-teacher", password="test-password", role=User.Role.TEACHER
        )
        self.school_class.class_teacher = teacher
        self.school_class.save(update_fields=["class_teacher"])
        requirement = Requirement.objects.filter(
            class_group=Requirement.ClassGroup.P4_P7, scholar_type="DAY"
        ).first()
        self.client.force_login(teacher)

        response = self.client.post(reverse("students:requirements_register"), {
            "class_id": self.school_class.pk,
            "term": "TERM1",
            "year": 2026,
            "scholar_type": "DAY",
            "brought": [f"{self.student.pk}:{requirement.pk}"],
        })

        self.assertEqual(response.status_code, 302)
        record = StudentRequirement.objects.get(
            student=self.student, requirement=requirement, term="TERM1", year=2026
        )
        self.assertTrue(record.brought)
        self.assertEqual(record.recorded_by, teacher)

    def test_teacher_cannot_access_another_class_requirements(self):
        teacher = User.objects.create_user(
            username="restricted-teacher", password="test-password", role=User.Role.TEACHER
        )
        assigned_class = SchoolClass.objects.create(name="P5", class_teacher=teacher)
        self.client.force_login(teacher)

        response = self.client.get(reverse("students:requirements_register"), {
            "class_id": self.school_class.pk, "term": "TERM1", "year": 2026,
        })
        export_response = self.client.get(reverse("students:export_requirements_register"), {
            "class_id": self.school_class.pk, "term": "TERM1", "year": 2026,
        })

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context["selected_class"])
        self.assertContains(response, assigned_class.name)
        self.assertNotContains(response, self.student.full_name)
        self.assertEqual(export_response.status_code, 404)

class StudentSearchTests(TestCase):
    def setUp(self):
        headteacher = User.objects.create_user(
            username="search-head", password="test-password", role=User.Role.HEADTEACHER
        )
        school_class = SchoolClass.objects.create(name="Search Class")
        self.jane = school_class.students.create(
            first_name="Jane", last_name="Nakato", gender="F", date_of_birth="2015-01-01",
            guardian_name="Guardian", guardian_phone="0700000000",
        )
        self.john = school_class.students.create(
            first_name="John", last_name="Kato", gender="M", date_of_birth="2015-02-01",
            guardian_name="Guardian", guardian_phone="0700000001",
        )
        self.client.force_login(headteacher)

    def test_searches_by_multi_part_name(self):
        response = self.client.get(reverse("students:headteacher_dashboard"), {"q": "jane nak"})

        self.assertContains(response, self.jane.full_name)
        self.assertNotContains(response, self.john.full_name)
        self.assertContains(response, 'value="jane nak"')

    def test_searches_by_partial_admission_number(self):
        response = self.client.get(
            reverse("students:headteacher_dashboard"), {"q": self.john.admission_number[-3:]}
        )

        self.assertContains(response, self.john.full_name)
        self.assertNotContains(response, self.jane.full_name)

    def test_displays_no_matches_message(self):
        response = self.client.get(reverse("students:headteacher_dashboard"), {"q": "not-a-pupil"})

        self.assertContains(response, 'No students match "not-a-pupil".')
