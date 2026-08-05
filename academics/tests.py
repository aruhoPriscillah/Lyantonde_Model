import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from accounts.models import User
from students.models import SchoolClass

from .forms import BulkResultFilterForm, ResultForm
from .models import Result, Subject
from .views import is_lower_primary_class, is_nursery_class, is_upper_primary_class, nursery_report_rows


class NurseryClassDetectionTests(SimpleTestCase):
    def test_supported_nursery_class_names(self):
        for name in ("Baby Class", "Nursery", "Nursery Class", "Middle", "Middle Class", "Top Class"):
            with self.subTest(name=name):
                self.assertTrue(is_nursery_class(SchoolClass(name=name)))

    def test_primary_class_is_not_nursery(self):
        self.assertFalse(is_nursery_class(SchoolClass(name="Primary Four")))


class ResultFormTests(SimpleTestCase):
    def test_remarks_is_a_dropdown_with_all_remark_choices(self):
        form = ResultForm()

        self.assertEqual(form.fields["remarks"].widget.__class__.__name__, "Select")
        self.assertEqual(
            list(form.fields["remarks"].choices),
            [
                ("", "---------"),
                ("POOR", "Poor"),
                ("FAIR", "Fair"),
                ("GOOD", "Good"),
                ("VERY_GOOD", "Very Good"),
                ("EXCELLENT", "Excellent"),
            ],
        )


class LowerPrimaryClassDetectionTests(SimpleTestCase):
    def test_supported_lower_primary_names(self):
        for name in ("P1", "P 2", "P3 Class", "Primary 1", "Primary Two", "Primary Three"):
            with self.subTest(name=name):
                self.assertTrue(is_lower_primary_class(SchoolClass(name=name)))

    def test_other_classes_are_not_lower_primary(self):
        for name in ("Baby Class", "P4", "Primary Four"):
            with self.subTest(name=name):
                self.assertFalse(is_lower_primary_class(SchoolClass(name=name)))


class UpperPrimaryClassDetectionTests(SimpleTestCase):
    def test_supported_upper_primary_names(self):
        for name in ("P4", "P 5", "P6 Class", "P7", "Primary 4", "Primary Five", "Primary Seven"):
            with self.subTest(name=name):
                self.assertTrue(is_upper_primary_class(SchoolClass(name=name)))

    def test_other_classes_are_not_upper_primary(self):
        for name in ("Top Class", "P3", "Primary Three"):
            with self.subTest(name=name):
                self.assertFalse(is_upper_primary_class(SchoolClass(name=name)))


class ClassSubjectDropdownTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        for name in (
            "Numbers", "English", "Reading", "Health Habits", "Social Development", "Drawing", "Writing",
            "Mathematics", "Literacy I", "Literacy II (Reading)", "Religious Education", "Luganda", "SST", "Science",
        ):
            Subject.objects.get_or_create(name=name)

    def subject_names(self, form):
        return set(form.fields["subject"].queryset.values_list("name", flat=True))

    def test_nursery_dropdown_only_has_nursery_subjects(self):
        form = ResultForm(teacher_class=SchoolClass.objects.create(name="Baby Class"))
        self.assertEqual(self.subject_names(form), {
            "Numbers", "English", "Reading", "Health Habits", "Social Development", "Drawing", "Writing",
        })

    def test_lower_primary_dropdown_only_has_lower_primary_subjects(self):
        form = ResultForm(teacher_class=SchoolClass.objects.create(name="P2"))
        self.assertEqual(self.subject_names(form), {
            "Mathematics", "English", "Literacy I", "Literacy II (Reading)", "Religious Education", "Luganda",
        })

    def test_upper_primary_single_and_bulk_dropdowns_match(self):
        school_class = SchoolClass.objects.create(name="Primary 6")
        expected = {"Mathematics", "English", "SST", "Science"}
        self.assertEqual(self.subject_names(ResultForm(teacher_class=school_class)), expected)
        self.assertEqual(self.subject_names(BulkResultFilterForm(teacher_class=school_class)), expected)

    def test_nursery_report_includes_all_subjects_without_results(self):
        school_class = SchoolClass.objects.create(name="Top Class")
        student = school_class.students.create(
            first_name="Test",
            last_name="Pupil",
            gender="F",
            date_of_birth="2021-01-01",
            guardian_name="Parent",
            guardian_phone="0700000000",
        )

        rows = nursery_report_rows(student, "TERM1", 2026)

        self.assertEqual([row["subject"] for row in rows], [
            "Numbers", "English", "Reading", "Health Habits", "Social Development", "Drawing", "Writing",
        ])
        self.assertTrue(all(row["score"] is None for row in rows))


class ExcelResultImportTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="excel-teacher", password="test-password", role=User.Role.TEACHER
        )
        self.school_class = SchoolClass.objects.create(name="P6", class_teacher=self.teacher)
        self.student = self.school_class.students.create(
            first_name="Excel",
            last_name="Pupil",
            gender="M",
            date_of_birth="2015-01-01",
            guardian_name="Parent",
            guardian_phone="0700000000",
        )
        self.math = Subject.objects.get(name="Mathematics")
        self.client.force_login(self.teacher)

    def workbook_upload(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Admission Number", "Pupil Name", "Subject", "Score", "Remarks", "Term", "Year"])
        for row in rows:
            sheet.append(row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            "results.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_download_template_contains_class_pupils_and_allowed_subjects(self):
        response = self.client.get(reverse("academics:download_results_template"), {"term": "TERM1", "year": 2026})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual({row[0] for row in rows}, {self.student.admission_number})
        self.assertEqual({row[2] for row in rows}, {"Mathematics", "English", "SST", "Science"})

    def test_imports_valid_excel_results(self):
        upload = self.workbook_upload([[
            self.student.admission_number, self.student.full_name, "Mathematics", 84, "Very Good", "TERM1", 2026,
        ]])

        response = self.client.post(reverse("academics:import_results_excel"), {"excel_file": upload})

        self.assertEqual(response.status_code, 302)
        result = Result.objects.get(student=self.student, subject=self.math, term="TERM1", year=2026)
        self.assertEqual(result.score, 84)
        self.assertEqual(result.remarks, "VERY_GOOD")

    def test_invalid_row_prevents_partial_import(self):
        upload = self.workbook_upload([
            [self.student.admission_number, self.student.full_name, "Mathematics", 84, "Good", "TERM1", 2026],
            [self.student.admission_number, self.student.full_name, "Luganda", 70, "Good", "TERM1", 2026],
        ])

        response = self.client.post(reverse("academics:import_results_excel"), {"excel_file": upload}, follow=True)

        self.assertContains(response, "subject is not allowed")
        self.assertFalse(Result.objects.filter(student=self.student).exists())
